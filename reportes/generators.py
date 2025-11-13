"""
Generadores de archivos para reportes en diferentes formatos.
Soporta Excel, PDF, CSV y JSON para exportar datos de reportes.
"""
import logging
import io
import csv
from datetime import datetime
from django.http import HttpResponse
from django.utils import timezone

logger = logging.getLogger(__name__)

# Imports opcionales con fallbacks
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl no disponible. Reportes Excel deshabilitados.")

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logger.warning("reportlab no disponible. Reportes PDF deshabilitados.")


class ExcelGenerator:
    """Generador de reportes en formato Excel."""
    
    def __init__(self):
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl no está instalado")
    
    def generate(self, data, headers: list, title: str = "Reporte") -> HttpResponse:
        """
        Genera un archivo Excel con los datos proporcionados.
        
        Args:
            data: QuerySet o lista de diccionarios con los datos
            headers: Lista de nombres de columnas
            title: Título del reporte
            
        Returns:
            HttpResponse: Respuesta con el archivo Excel
        """
        logger.info(f"Generando reporte Excel: {title}")
        
        # Crear workbook y worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Reporte"
        
        # Estilo para el título
        title_font = Font(bold=True, size=16, color="FFFFFF")
        title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        # Estilo para headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Bordes
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Escribir título
        worksheet.merge_cells(f'A1:{get_column_letter(len(headers))}1')
        worksheet['A1'] = title
        worksheet['A1'].font = title_font
        worksheet['A1'].fill = title_fill
        worksheet['A1'].alignment = title_alignment
        worksheet.row_dimensions[1].height = 30
        
        # Escribir fecha de generación
        fecha_gen = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        worksheet.merge_cells(f'A2:{get_column_letter(len(headers))}2')
        worksheet['A2'] = fecha_gen
        worksheet['A2'].alignment = Alignment(horizontal="right")
        
        # Escribir headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Escribir datos
        row = 5
        for item in data:
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=row, column=col)
                
                # Extraer valor del item
                if isinstance(item, dict):
                    # Para valores de QuerySet.values()
                    value = list(item.values())[col-1] if col-1 < len(item.values()) else ''
                else:
                    # Para objetos modelo
                    value = getattr(item, header.lower().replace(' ', '_'), '')
                
                # Formatear valor
                if isinstance(value, datetime):
                    cell.value = value.strftime('%d/%m/%Y %H:%M')
                elif isinstance(value, bool):
                    cell.value = "Sí" if value else "No"
                elif value is None:
                    cell.value = ""
                else:
                    cell.value = str(value)
                
                cell.border = thin_border
                
                # Alineación según tipo de dato
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
            row += 1
        
        # Ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Crear respuesta HTTP
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f"{title.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"Reporte Excel generado: {filename}")
        return response


class PDFGenerator:
    """Generador de reportes en formato PDF."""
    
    def __init__(self):
        if not PDF_AVAILABLE:
            raise ImportError("reportlab no está instalado")
    
    def generate(self, data, headers: list, title: str = "Reporte") -> HttpResponse:
        """
        Genera un archivo PDF con los datos proporcionados.
        
        Args:
            data: QuerySet o lista de diccionarios con los datos
            headers: Lista de nombres de columnas
            title: Título del reporte
            
        Returns:
            HttpResponse: Respuesta con el archivo PDF
        """
        logger.info(f"Generando reporte PDF: {title}")
        
        # Crear buffer
        buffer = io.BytesIO()
        
        # Crear documento PDF
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Centrado
        )
        
        # Elementos del documento
        elements = []
        
        # Título
        elements.append(Paragraph(title, title_style))
        
        # Fecha de generación
        fecha = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        elements.append(Paragraph(fecha, styles['Normal']))
        elements.append(Spacer(1, 12))
        
        # Preparar datos para la tabla
        table_data = [headers]  # Primera fila son los headers
        
        for item in data:
            row = []
            for col, header in enumerate(headers):
                if isinstance(item, dict):
                    value = list(item.values())[col] if col < len(item.values()) else ''
                else:
                    value = getattr(item, header.lower().replace(' ', '_'), '')
                
                # Formatear valor
                if isinstance(value, datetime):
                    formatted_value = value.strftime('%d/%m/%Y')
                elif isinstance(value, bool):
                    formatted_value = "Sí" if value else "No"
                elif value is None:
                    formatted_value = ""
                else:
                    formatted_value = str(value)[:50]  # Truncar textos largos
                
                row.append(formatted_value)
            table_data.append(row)
        
        # Crear tabla
        table = Table(table_data)
        
        # Estilo de la tabla
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])
        
        table.setStyle(table_style)
        elements.append(table)
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        filename = f"{title.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"Reporte PDF generado: {filename}")
        return response


class CSVGenerator:
    """Generador de reportes en formato CSV."""
    
    def generate(self, data, headers: list, title: str = "Reporte") -> HttpResponse:
        """
        Genera un archivo CSV con los datos proporcionados.
        
        Args:
            data: QuerySet o lista de diccionarios con los datos
            headers: Lista de nombres de columnas
            title: Título del reporte
            
        Returns:
            HttpResponse: Respuesta con el archivo CSV
        """
        logger.info(f"Generando reporte CSV: {title}")
        
        # Crear buffer de texto
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Escribir título y fecha
        writer.writerow([f"# {title}"])
        writer.writerow([f"# Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')}"])
        writer.writerow([])  # Línea vacía
        
        # Escribir headers
        writer.writerow(headers)
        
        # Escribir datos
        for item in data:
            row = []
            for col, header in enumerate(headers):
                if isinstance(item, dict):
                    value = list(item.values())[col] if col < len(item.values()) else ''
                else:
                    value = getattr(item, header.lower().replace(' ', '_'), '')
                
                # Formatear valor
                if isinstance(value, datetime):
                    formatted_value = value.strftime('%d/%m/%Y %H:%M')
                elif isinstance(value, bool):
                    formatted_value = "Sí" if value else "No"
                elif value is None:
                    formatted_value = ""
                else:
                    formatted_value = str(value)
                
                row.append(formatted_value)
            writer.writerow(row)
        
        # Crear respuesta HTTP
        response = HttpResponse(
            output.getvalue(),
            content_type='text/csv; charset=utf-8-sig'
        )
        
        filename = f"{title.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"Reporte CSV generado: {filename}")
        return response


def generate_report(data, headers: list, format_type: str, title: str = "Reporte") -> HttpResponse:
    """
    Función de conveniencia para generar reportes en diferentes formatos.
    
    Args:
        data: Datos del reporte
        headers: Encabezados de columnas
        format_type: Tipo de formato (excel, pdf, csv)
        title: Título del reporte
        
    Returns:
        HttpResponse: Respuesta con el archivo generado
    """
    format_type = format_type.lower()
    
    if format_type == 'excel':
        if not EXCEL_AVAILABLE:
            raise ImportError("Excel no disponible. Instale openpyxl.")
        generator = ExcelGenerator()
        return generator.generate(data, headers, title)
    
    elif format_type == 'pdf':
        if not PDF_AVAILABLE:
            raise ImportError("PDF no disponible. Instale reportlab.")
        generator = PDFGenerator()
        return generator.generate(data, headers, title)
    
    elif format_type == 'csv':
        generator = CSVGenerator()
        return generator.generate(data, headers, title)
    
    else:
        raise ValueError(f"Formato '{format_type}' no soportado. Use: excel, pdf, csv")